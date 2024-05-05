#——————————————————————————————————————————读取xls文件————————————————————————————————————————#
#————————————————————————————————————————————常量————————————————————————————————————————————#



#———————————————————————————————————————————头文件————————————————————————————————————————————#

import os
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from Cell import *

#—————————————————————————————————————————————类—————————————————————————————————————————————#

class XlsReader():
    '''读取xls文件并存储的类'''

    def __init__(self, file_name: str, data_only=True) -> None:
        '''初始化类
            file_name: 要读取的文件名称
            data_only: 是否只读取数据'''
        self.file_name = file_name
        self.data = {}
        self.data_only = data_only
        self.read_file()
    
    def read_file(self):
        '''读取表格文件, 并存储在data中'''
        workboot = openpyxl.load_workbook(self.file_name, data_only=self.data_only)
        for sheetname in workboot.sheetnames:
            worksheet = workboot[sheetname]
            self.data[sheetname] = self.read_one_sheet(worksheet=worksheet)
            
    
    def read_one_sheet(self, worksheet: Worksheet) -> list:
        '''读取一个工作簿'''
        merge_ranges = worksheet.merged_cells.ranges
        r_idx, c_idx = 0, 0
        rt = []
        # 列举出所有的单元格
        merge_cells = {}
        for merge_range in merge_ranges:
            for row in range(merge_range.min_row - 1, merge_range.max_row):
                for col in range(merge_range.min_col - 1, merge_range.max_col):
                    merge_cells[f"[{row}, {col}]"] = {"lt_pos": [merge_range.min_row - 1, merge_range.min_col - 1],
                                                      "shape": [merge_range.max_row - merge_range.min_row + 1, merge_range.max_col - merge_range.min_col + 1]}
        for row in worksheet:
            temp = []
            for cell in row:
                if not merge_cells.get(f"[{r_idx}, {c_idx}]"):
                    temp_cell = BaseCell(cell.value, pos=[r_idx, c_idx])
                else:
                    temp_cell = MergeCell(cell.value, pos=[r_idx, c_idx], lt_pos=merge_cells[f"[{r_idx}, {c_idx}]"]["lt_pos"],
                                          shape=merge_cells[f"[{r_idx}, {c_idx}]"]["shape"])                    
                temp.append(temp_cell)
                c_idx += 1
            r_idx += 1
            c_idx = 0
            rt.append(temp)
        return rt
        


#————————————————————————————————————————————程序————————————————————————————————————————————#

# reader = XlsReader("/home/tianxj/myCode/useful_program/xls_to_latex_label/xlsx_file/test.xlsx", data_only=True)
# for row in reader.data["工作表1"]:
#     for cell in row:
#         print(cell)